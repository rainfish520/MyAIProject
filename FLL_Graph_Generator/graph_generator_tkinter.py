# -*- coding: utf-8 -*-
"""
Graph文件生成器 - GUI版本
使用tkinter构建，用于生成游戏动画状态机.graph文件

版本: 4.0 (v4)
支持: 多技能、三段式结构、瞬发式结构、V2批量导出、自动填表

更新:
- 增加Graph名称输入框
- 技能名称来自节点配置的输入框
- 节点配置增加结构类型下拉框
- 移除位置输入框，位置自动排列
- V2批量导出配置（场所类型/车辆类型）
- 自动填表功能（根据模型编号查询动作数据）
- 导出graph并自动填表功能
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import tkinter.simpledialog
import sys
import os
import pandas as pd
import xml.etree.ElementTree as ET
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# 导入我们的模块
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from models import GraphModel, GraphLayer, GraphNode, StructureType, CarType, Event, Cue
from serializer import GraphSerializer


class GraphGeneratorApp:
    """Graph生成器主应用程序"""
    
    def __init__(self, root):
        self.root = root
        self.root.title("Graph文件生成器 v4.0 - 动画状态机配置工具")
        self.root.geometry("900x700")
        self.root.resizable(True, True)
        
        # 最小窗口大小
        self.root.minsize(800, 600)
        
        # 初始化数据模型
        self.model = GraphModel()
        
        # 当前选中的层索引
        self.current_layer_index = 0
        
        # 样式配置
        self.setup_styles()
        
        # 创建界面
        self.setup_ui()
        
        # 初始化默认图层
        self.model.add_layer(StructureType.THREE_STAGE)
        self.model.layers[0].name = "大招加速"
        self.refresh_layer_list()
        self.load_layer_to_ui(0)
        
        self.log("Graph生成器 v4.0 启动成功")
        self.log("支持三段式和瞬发式结构")
    
    def setup_styles(self):
        """配置界面样式"""
        self.style = ttk.Style()
        self.style.theme_use('vista')
        
        # 自定义样式
        self.style.configure("Title.TLabel", font=("微软雅黑", 14, "bold"), foreground="#2c3e50")
        self.style.configure("Section.TLabelframe", font=("微软雅黑", 10, "bold"))
        self.style.configure("Section.TLabel", font=("微软雅黑", 9))
        self.style.configure("Action.TButton", font=("微软雅黑", 10), padding=6)
        self.style.configure("Log.TText", font=("Consolas", 9))
    
    def setup_ui(self):
        """创建用户界面"""
        # 主容器 - 使用PanedWindow左右分割
        main_paned = ttk.PanedWindow(self.root, orient=tk.HORIZONTAL)
        main_paned.pack(fill=tk.BOTH, expand=True)
        
        # 左侧面板 - 技能管理
        left_frame = ttk.Frame(main_paned, width=200)
        main_paned.add(left_frame, weight=1)
        self.create_left_panel(left_frame)
        
        # 右侧面板 - 工作区
        right_frame = ttk.Frame(main_paned)
        main_paned.add(right_frame, weight=4)
        self.create_right_panel(right_frame)
        
        # 底部状态栏
        self.create_status_bar()
    
    def create_left_panel(self, parent):
        """创建左侧技能管理面板"""
        frame = ttk.LabelFrame(parent, text=" 技能列表 ", padding="10")
        frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # Graph名称输入
        name_frame = ttk.Frame(frame)
        name_frame.pack(fill=tk.X, pady=(0, 10))
        ttk.Label(name_frame, text="Graph名称:").pack(anchor=tk.W)
        self.entry_graph_name = ttk.Entry(name_frame, width=25)
        self.entry_graph_name.pack(fill=tk.X, pady=(2, 0))
        self.entry_graph_name.insert(0, "生成模版")
        self.entry_graph_name.bind('<<KeyRelease>>', self.on_graph_name_change)
        
        # 模型信息输入（用于自动填表）
        model_info_frame = ttk.LabelFrame(frame, text=" 模型信息 ", padding="5")
        model_info_frame.pack(fill=tk.X, pady=(0, 10))
        
        # 模型编号
        ttk.Label(model_info_frame, text="模型编号:").grid(row=0, column=0, sticky=tk.W, pady=2)
        self.entry_model_number = ttk.Entry(model_info_frame, width=20)
        self.entry_model_number.grid(row=0, column=1, sticky=tk.W, pady=2)
        self.entry_model_number.insert(0, "1000")
        
        # 模型备注
        ttk.Label(model_info_frame, text="模型备注:").grid(row=1, column=0, sticky=tk.W, pady=2)
        self.entry_model_remark = ttk.Entry(model_info_frame, width=20)
        self.entry_model_remark.grid(row=1, column=1, sticky=tk.W, pady=2)
        self.entry_model_remark.insert(0, "备注")
        
        # 技能列表
        list_frame = ttk.Frame(frame)
        list_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 10))
        
        ttk.Label(list_frame, text="技能列表:").pack(anchor=tk.W, pady=(0, 5))
        
        self.layer_listbox = tk.Listbox(list_frame, height=12, selectmode=tk.SINGLE)
        self.layer_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self.layer_listbox.bind('<<ListboxSelect>>', self.on_layer_select)
        
        scrollbar = ttk.Scrollbar(list_frame, orient=tk.VERTICAL, command=self.layer_listbox.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.layer_listbox.config(yscrollcommand=scrollbar.set)
        
        # 按钮区域
        btn_frame = ttk.Frame(frame)
        btn_frame.pack(fill=tk.X)
        
        ttk.Button(btn_frame, text="添加技能", command=self.add_layer).pack(fill=tk.X, pady=2)
        ttk.Button(btn_frame, text="删除技能", command=self.delete_layer).pack(fill=tk.X, pady=(10, 2))
    
    def create_right_panel(self, parent):
        """创建右侧工作区"""
        # 标题
        title_frame = ttk.Frame(parent, padding="10")
        title_frame.pack(fill=tk.X)
        ttk.Label(title_frame, text="Graph文件生成器", style="Title.TLabel").pack(side=tk.LEFT)
        
        # Notebook用于多标签页
        notebook = ttk.Notebook(parent)
        notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        
        # 标签页1: 节点配置
        self.node_config_frame = ttk.Frame(notebook, padding="10")
        notebook.add(self.node_config_frame, text=" 节点配置 ")
        self.create_node_config_tab()
        
        # 标签页2: 事件配置
        self.event_config_frame = ttk.Frame(notebook, padding="10")
        notebook.add(self.event_config_frame, text=" 事件配置 ")
        self.create_event_config_tab()
        
        # 标签页3: 导出
        self.export_frame = ttk.Frame(notebook, padding="10")
        notebook.add(self.export_frame, text=" 导出 ")
        self.create_export_tab()
        
        # 标签页4: 涂装生成
        self.paint_frame = ttk.Frame(notebook, padding="10")
        notebook.add(self.paint_frame, text=" 涂装生成 ")
        self.create_paint_tab()
        
        # 标签页5: 预览
        self.preview_frame = ttk.Frame(notebook, padding="10")
        notebook.add(self.preview_frame, text=" 预览 ")
        self.create_preview_tab()
    
    def create_node_config_tab(self):
        """创建节点配置标签页"""
        # 当前技能信息
        info_frame = ttk.LabelFrame(self.node_config_frame, text=" 当前技能配置 ", padding="10")
        info_frame.pack(fill=tk.X, pady=(0, 10))
        
        row = 0
        # 技能名称输入
        ttk.Label(info_frame, text="技能名称:").grid(row=row, column=0, sticky=tk.E, padx=5, pady=5)
        self.entry_skill_name = ttk.Entry(info_frame, width=30)
        self.entry_skill_name.grid(row=row, column=1, sticky=tk.W, padx=5, pady=5)
        # 绑定两个事件：按键释放和失去焦点
        self.entry_skill_name.bind('<<KeyRelease>>', self.on_skill_name_change)
        self.entry_skill_name.bind('<FocusOut>', self.on_skill_name_change)
        # 技能名称变化时也更新文件名预览
        self.entry_skill_name.bind('<<KeyRelease>>', lambda e: self.update_filename_preview())
        self.entry_skill_name.bind('<FocusOut>', lambda e: self.update_filename_preview())
        
        row += 1
        # 结构类型下拉框
        ttk.Label(info_frame, text="结构类型:").grid(row=row, column=0, sticky=tk.E, padx=5, pady=5)
        self.structure_combo = ttk.Combobox(info_frame, state="readonly", width=28)
        self.structure_combo['values'] = ("三段式 (Start-Loop-End)", "瞬发式")
        self.structure_combo.grid(row=row, column=1, sticky=tk.W, padx=5, pady=5)
        self.structure_combo.current(0)
        self.structure_combo.bind('<<ComboboxSelected>>', self.on_structure_change)
        
        # 场所类型（复选）- 每个技能单独设置
        row += 1
        ttk.Label(info_frame, text="车辆类型:").grid(row=row, column=0, sticky=tk.E, padx=5, pady=5)
        car_type_frame = ttk.Frame(info_frame)
        car_type_frame.grid(row=row, column=1, sticky=tk.W, padx=5, pady=5)
        
        self.var_car_type = tk.StringVar(value="NEW_CAR")
        ttk.Radiobutton(car_type_frame, text="新车", variable=self.var_car_type, value="NEW_CAR").pack(side=tk.LEFT, padx=5)
        ttk.Radiobutton(car_type_frame, text="老车", variable=self.var_car_type, value="OLD_CAR").pack(side=tk.LEFT, padx=5)
        
        # 绑定车辆类型变更事件
        self.trace_car_type_id = self.var_car_type.trace('w', self.on_place_change)
        
        row += 1
        ttk.Label(info_frame, text="场所类型:").grid(row=row, column=0, sticky=tk.E, padx=5, pady=5)
        place_frame = ttk.Frame(info_frame)
        place_frame.grid(row=row, column=1, sticky=tk.W, padx=5, pady=5)
        
        # 展厅选项（车身/底盘复选，可取消勾选）
        ttk.Label(place_frame, text="展厅:").grid(row=0, column=0, sticky=tk.W, padx=2, pady=2)
        self.var_showroom_body = tk.BooleanVar(value=True)
        ttk.Checkbutton(place_frame, text="车身", variable=self.var_showroom_body).grid(row=0, column=1, sticky=tk.W, padx=2)
        self.var_showroom_chassis = tk.BooleanVar(value=False)
        ttk.Checkbutton(place_frame, text="底盘", variable=self.var_showroom_chassis).grid(row=0, column=2, sticky=tk.W, padx=2)
        
        # 赛道选项（车身/底盘复选，可取消勾选）
        ttk.Label(place_frame, text="赛道:").grid(row=1, column=0, sticky=tk.W, padx=2, pady=2)
        self.var_race_body = tk.BooleanVar(value=True)
        ttk.Checkbutton(place_frame, text="车身", variable=self.var_race_body).grid(row=1, column=1, sticky=tk.W, padx=2)
        self.var_race_chassis = tk.BooleanVar(value=False)
        ttk.Checkbutton(place_frame, text="底盘", variable=self.var_race_chassis).grid(row=1, column=2, sticky=tk.W, padx=2)
        
        # 绑定场所变更事件并保存trace ID
        self.trace_showroom_id = self.var_showroom_body.trace('w', self.on_place_change)
        self.trace_showroom_chassis_id = self.var_showroom_chassis.trace('w', self.on_place_change)
        self.trace_race_id = self.var_race_body.trace('w', self.on_place_change)
        self.trace_race_chassis_id = self.var_race_chassis.trace('w', self.on_place_change)
        
        # 节点配置
        nodes_frame = ttk.LabelFrame(self.node_config_frame, text=" 节点配置 ", padding="10")
        nodes_frame.pack(fill=tk.BOTH, expand=True)
        
        # 节点选择
        select_frame = ttk.Frame(nodes_frame)
        select_frame.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(select_frame, text="选择节点:").pack(side=tk.LEFT, padx=5)
        self.node_combo = ttk.Combobox(select_frame, state="readonly", width=25)
        self.node_combo.pack(side=tk.LEFT, padx=5)
        self.node_combo.bind('<<ComboboxSelected>>', self.on_node_select)
        
        # 节点属性表单
        self.node_form_frame = ttk.LabelFrame(nodes_frame, text=" 节点属性 ", padding="10")
        self.node_form_frame.pack(fill=tk.BOTH, expand=True)
        
        # 动态创建表单
        self.create_node_form()
    
    def create_node_form(self):
        """创建节点属性表单 - 简化版，移除位置输入"""
        # 清除旧控件
        for widget in self.node_form_frame.winfo_children():
            widget.destroy()
        
        # 动画名称
        row = 0
        ttk.Label(self.node_form_frame, text="动画名称:").grid(row=row, column=0, sticky=tk.E, padx=5, pady=5)
        self.entry_anim_name = ttk.Entry(self.node_form_frame, width=30)
        self.entry_anim_name.grid(row=row, column=1, sticky=tk.W, padx=5, pady=5)
        self.entry_anim_name.insert(0, "no_anim")
        
        # 引用时间
        row += 1
        ttk.Label(self.node_form_frame, text="引用时间:").grid(row=row, column=0, sticky=tk.E, padx=5, pady=5)
        self.spin_add_ref = ttk.Spinbox(self.node_form_frame, from_=0, to=10, width=28)
        self.spin_add_ref.set(0)
        self.spin_add_ref.grid(row=row, column=1, sticky=tk.W, padx=5, pady=5)
        
        # 单次播放
        row += 1
        self.var_single_play = tk.BooleanVar(value=True)
        ttk.Checkbutton(self.node_form_frame, text="单次播放", variable=self.var_single_play).grid(
            row=row, column=0, columnspan=2, sticky=tk.W, padx=5, pady=5)
        
        # 应用按钮
        row += 1
        ttk.Button(self.node_form_frame, text="应用更改", command=self.apply_node_changes).grid(
            row=row, column=0, columnspan=2, pady=15)
    
    def create_event_config_tab(self):
        """创建事件配置标签页"""
        # 说明
        info_frame = ttk.LabelFrame(self.event_config_frame, text=" 事件说明 ", padding="10")
        info_frame.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(info_frame, text="事件用于触发状态转换", foreground="#7f8c8d").pack(anchor=tk.W)
        ttk.Label(info_frame, text="每个节点可以包含多个事件", foreground="#7f8c8d").pack(anchor=tk.W)
        
        # 当前节点事件
        events_frame = ttk.LabelFrame(self.event_config_frame, text=" 当前节点事件 ", padding="10")
        events_frame.pack(fill=tk.BOTH, expand=True)
        
        # 事件列表框
        list_frame = ttk.Frame(events_frame)
        list_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 10))
        
        self.event_listbox = tk.Listbox(list_frame, height=10)
        self.event_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        event_scroll = ttk.Scrollbar(list_frame, orient=tk.VERTICAL, command=self.event_listbox.yview)
        event_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        self.event_listbox.config(yscrollcommand=event_scroll.set)
        
        # 事件按钮
        btn_frame = ttk.Frame(events_frame)
        btn_frame.pack(fill=tk.X)
        
        ttk.Button(btn_frame, text="添加事件", command=self.add_event).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="删除事件", command=self.delete_event).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="刷新列表", command=self.refresh_event_list).pack(side=tk.RIGHT, padx=5)
    
    def create_export_tab(self):
        """创建导出标签页"""
        # V2导出配置 - 场所和车辆类型（现在由每个技能单独配置）
        v2_config_frame = ttk.LabelFrame(self.export_frame, text=" V2批量导出配置 ", padding="15")
        v2_config_frame.pack(fill=tk.X, pady=(0, 10))
        
        # 说明文字
        ttk.Label(v2_config_frame, text="场所类型（展厅/赛道）和车辆类型（车身/底盘）在每个技能设置中单独配置",
                  foreground="#7f8c8d").pack(pady=5)
        
        # 文件名预览
        name_preview_frame = ttk.Frame(v2_config_frame)
        name_preview_frame.pack(fill=tk.X, pady=5)
        ttk.Label(name_preview_frame, text="文件名预览:").pack(side=tk.LEFT, padx=5)
        self.label_filename_preview = ttk.Label(name_preview_frame, text="(根据技能设置自动生成)", foreground="#7f8c8d")
        self.label_filename_preview.pack(side=tk.LEFT, padx=5)
        
        # 分隔线
        ttk.Separator(self.export_frame, orient=tk.HORIZONTAL).pack(fill=tk.X, pady=5)
        
        # 自动填表功能
        autofill_frame = ttk.LabelFrame(self.export_frame, text=" 自动填表功能 ", padding="15")
        autofill_frame.pack(fill=tk.X, pady=(0, 10))
        
        # 模型编号输入
        model_input_frame = ttk.Frame(autofill_frame)
        model_input_frame.pack(fill=tk.X, pady=5)
        ttk.Label(model_input_frame, text="模型编号:").pack(side=tk.LEFT, padx=5)
        self.entry_model_id = ttk.Entry(model_input_frame, width=20)
        self.entry_model_id.pack(side=tk.LEFT, padx=5)
        ttk.Label(model_input_frame, text="(输入动作组合中的编号)", foreground="#7f8c8d").pack(side=tk.LEFT, padx=5)
        
        # 查找按钮
        btn_frame = ttk.Frame(autofill_frame)
        btn_frame.pack(fill=tk.X, pady=5)
        ttk.Button(btn_frame, text="查询动作数据", command=self.query_model_actions).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="导出到Excel", command=self.export_to_excel).pack(side=tk.LEFT, padx=5)
        
        # 结果显示
        result_frame = ttk.Frame(autofill_frame)
        result_frame.pack(fill=tk.BOTH, expand=True, pady=5)
        
        # 结果文本框
        self.result_text = tk.Text(result_frame, height=8, font=("Consolas", 9))
        self.result_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        result_scroll = ttk.Scrollbar(result_frame, orient=tk.VERTICAL, command=self.result_text.yview)
        result_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        self.result_text.config(yscrollcommand=result_scroll.set)
        
        # 分隔线
        ttk.Separator(self.export_frame, orient=tk.HORIZONTAL).pack(fill=tk.X, pady=5)
        
        # 导出选项
        options_frame = ttk.LabelFrame(self.export_frame, text=" 导出选项 ", padding="15")
        options_frame.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(options_frame, text="输出文件:").grid(row=0, column=0, sticky=tk.E, padx=5, pady=10)
        
        self.entry_output_path = ttk.Entry(options_frame, width=40)
        self.entry_output_path.grid(row=0, column=1, sticky=tk.EW, padx=5, pady=10)
        self.entry_output_path.insert(0, "E:/H65/Package/Char/graph/car_parts")
        
        ttk.Button(options_frame, text="浏览...", command=self.browse_output).grid(
            row=0, column=2, padx=5, pady=10)
        
        options_frame.columnconfigure(1, weight=1)
        
        # 导出按钮
        btn_frame = ttk.Frame(self.export_frame)
        btn_frame.pack(fill=tk.X, pady=10)
        
        ttk.Button(btn_frame, text="导出.graph文件", style="Action.TButton", 
                   command=self.export_graph).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="测试生成", 
                   command=self.test_export).pack(side=tk.LEFT, padx=5)
        
        # 日志
        log_frame = ttk.LabelFrame(self.export_frame, text=" 执行日志 ", padding="10")
        log_frame.pack(fill=tk.BOTH, expand=True)
        
        self.log_text = tk.Text(log_frame, height=12, state='disabled', 
                                bg="#1e1e1e", fg="#00ff00", font=("Consolas", 9))
        self.log_text.pack(fill=tk.BOTH, expand=True)
    
    def create_paint_tab(self):
        """创建涂装生成标签页"""
        # 输入配置
        input_frame = ttk.LabelFrame(self.paint_frame, text=" 涂装生成配置 ", padding="15")
        input_frame.pack(fill=tk.X, pady=(0, 10))
        
        # 模型动作编号输入（多个5位编号，连续输入）
        ttk.Label(input_frame, text="模型动作编号:").grid(row=0, column=0, sticky=tk.E, padx=5, pady=5)
        self.entry_old_model_id = ttk.Entry(input_frame, width=30)
        self.entry_old_model_id.grid(row=0, column=1, sticky=tk.W, padx=5, pady=5)
        ttk.Label(input_frame, text="(连续输入多个5位编号，如2600126002)", foreground="#7f8c8d").grid(row=0, column=2, sticky=tk.W, padx=5)
        
        # 新涂装名称输入
        ttk.Label(input_frame, text="新涂装名称:").grid(row=1, column=0, sticky=tk.E, padx=5, pady=5)
        self.entry_paint_name = ttk.Entry(input_frame, width=20)
        self.entry_paint_name.grid(row=1, column=1, sticky=tk.W, padx=5, pady=5)
        
        # 源目录
        ttk.Label(input_frame, text="源目录:").grid(row=2, column=0, sticky=tk.E, padx=5, pady=5)
        self.entry_paint_source_dir = ttk.Entry(input_frame, width=40)
        self.entry_paint_source_dir.grid(row=2, column=1, sticky=tk.EW, padx=5, pady=5)
        self.entry_paint_source_dir.insert(0, "E:/H65/Package/Char/graph/car_parts")
        ttk.Button(input_frame, text="浏览...", command=self.browse_paint_source_dir).grid(row=2, column=2, padx=5)
        
        # 输出目录
        ttk.Label(input_frame, text="输出目录:").grid(row=3, column=0, sticky=tk.E, padx=5, pady=5)
        self.entry_paint_output_dir = ttk.Entry(input_frame, width=40)
        self.entry_paint_output_dir.grid(row=3, column=1, sticky=tk.EW, padx=5, pady=5)
        self.entry_paint_output_dir.insert(0, "E:/H65/Package/Char/graph/car_parts")
        ttk.Button(input_frame, text="浏览...", command=self.browse_paint_output_dir).grid(row=3, column=2, padx=5)
        
        input_frame.columnconfigure(1, weight=1)
        
        # 模型信息
        ttk.Label(input_frame, text="模型编号:").grid(row=4, column=0, sticky=tk.E, padx=5, pady=5)
        self.entry_paint_model_number = ttk.Entry(input_frame, width=30)
        self.entry_paint_model_number.grid(row=4, column=1, sticky=tk.EW, padx=5, pady=5)
        
        ttk.Label(input_frame, text="模型备注:").grid(row=5, column=0, sticky=tk.E, padx=5, pady=5)
        self.entry_paint_model_remark = ttk.Entry(input_frame, width=30)
        self.entry_paint_model_remark.grid(row=5, column=1, sticky=tk.EW, padx=5, pady=5)
        
        # 生成按钮
        btn_frame = ttk.Frame(self.paint_frame)
        btn_frame.pack(fill=tk.X, pady=10)
        
        ttk.Button(btn_frame, text="生成涂装.graph文件", style="Action.TButton", 
                   command=self.generate_paint_graph).pack(side=tk.LEFT, padx=5)
        
        # 日志
        log_frame = ttk.LabelFrame(self.paint_frame, text=" 执行日志 ", padding="10")
        log_frame.pack(fill=tk.BOTH, expand=True)
        
        self.paint_log_text = tk.Text(log_frame, height=12, state='disabled', 
                                bg="#1e1e1e", fg="#00ff00", font=("Consolas", 9))
        self.paint_log_text.pack(fill=tk.BOTH, expand=True)
    
    def paint_log(self, message):
        """涂装生成日志"""
        self.paint_log_text.config(state='normal')
        self.paint_log_text.insert(tk.END, message + "\n")
        self.paint_log_text.see(tk.END)
        self.paint_log_text.config(state='disabled')
    
    def browse_paint_source_dir(self):
        """浏览源目录"""
        dirname = filedialog.askdirectory(title="选择源目录", initialdir=self.entry_paint_source_dir.get())
        if dirname:
            self.entry_paint_source_dir.delete(0, tk.END)
            self.entry_paint_source_dir.insert(0, dirname)
    
    def browse_paint_output_dir(self):
        """浏览输出目录"""
        dirname = filedialog.askdirectory(title="选择输出目录", initialdir=self.entry_paint_output_dir.get())
        if dirname:
            self.entry_paint_output_dir.delete(0, tk.END)
            self.entry_paint_output_dir.insert(0, dirname)
    
    def generate_paint_graph(self):
        """生成涂装graph文件"""
        old_ids_str = self.entry_old_model_id.get().strip()
        paint_name = self.entry_paint_name.get().strip()
        source_dir = self.entry_paint_source_dir.get().strip()
        output_dir = self.entry_paint_output_dir.get().strip()
        model_number = self.entry_paint_model_number.get().strip()
        model_remark = self.entry_paint_model_remark.get().strip()
        
        if not old_ids_str:
            messagebox.showwarning("警告", "请输入模型动作编号")
            return
        
        if not paint_name:
            messagebox.showwarning("警告", "请输入新涂装名称")
            return
        
        if not source_dir or not output_dir:
            messagebox.showwarning("警告", "请选择源目录和输出目录")
            return
        
        # 解析老车编号（连续输入多个5位编号，自动分割）
        try:
            # 去除所有空白字符
            clean_input = old_ids_str.replace(" ", "").replace("\n", "").replace("\t", "")
            
            # 验证输入是否为数字
            if not clean_input.isdigit():
                messagebox.showwarning("警告", "编号必须是数字")
                return
            
            # 验证长度是否为5的倍数
            if len(clean_input) % 5 != 0:
                messagebox.showwarning("警告", f"输入长度无效。当前字符数: {len(clean_input)}，请确保是5的倍数")
                return
            
            # 自动分割为5位编号列表
            old_ids = [clean_input[i:i+5] for i in range(0, len(clean_input), 5)]
            
            # 验证每个都是有效的5位编号
            for oid in old_ids:
                if int(oid) < 10000 or int(oid) > 99999:
                    messagebox.showwarning("警告", f"编号 {oid} 不是有效的5位编号")
                    return
            
        except Exception as e:
            messagebox.showwarning("警告", f"编号解析失败: {str(e)}")
            return
        
        self.paint_log(f"开始生成涂装...")
        self.paint_log(f"老车编号: {old_ids}")
        self.paint_log(f"新涂装名称: {paint_name}")
        
        # 读取Excel获取老车的graph信息
        try:
            excel_path = r"E:\H65\Doc\X_导入数据\01_美术资源\模型动作.xlsx"
            xl = pd.ExcelFile(excel_path)
            df_model = pd.read_excel(xl, '模型动作')
            
            # 查找老车编号对应的记录
            # 将old_ids转换为整数列表进行查询
            old_ids_int = [int(oid) for oid in old_ids]
            old_records = df_model[df_model['编号'].isin(old_ids_int)]
            
            if len(old_records) == 0:
                messagebox.showwarning("警告", f"未找到编号 {old_ids} 对应的模型动作记录")
                return
            
            self.paint_log(f"找到 {len(old_records)} 条老车记录")
            
            # 生成的记录
            model_rows = []
            vehicle_rows = []
            graph_files_created = []
            
            # 获取最大编号
            df_model['编号_str'] = df_model['编号'].astype(str)
            df_model_5digit = df_model[df_model['编号_str'].str.len() == 5]
            df_model_2xxx = df_model_5digit[df_model_5digit['编号'] >= 26000]
            max_model_id = int(df_model_2xxx['编号'].iloc[-1]) if len(df_model_2xxx) > 0 else 26000
            
            df_vehicle = pd.read_excel(xl, '车辆动作')
            max_vehicle_id = df_vehicle['编号'].max() if len(df_vehicle) > 0 else 0
            if pd.isna(max_vehicle_id):
                max_vehicle_id = 0
            vehicle_id = int(max_vehicle_id) + 1 if max_vehicle_id else 1
            
            for idx, record in old_records.iterrows():
                old_graph_path = record['GRAPH路径']
                old_remark = record['备注']
                
                # 从路径提取原graph文件名
                old_filename = os.path.basename(old_graph_path) if old_graph_path else ""
                
                if not old_filename:
                    self.paint_log(f"警告: 编号 {record['编号']} 没有GRAPH路径")
                    continue
                
                # 生成新文件名：在第一个下划线后插入涂装名称
                # 例如: template_base_show.graph -> template涂装名_base_show.graph
                parts = old_filename.split('_', 1)
                if len(parts) >= 2:
                    new_filename = f"{parts[0]}{paint_name}_{parts[1]}"
                else:
                    new_filename = f"{old_filename}"
                
                # 源文件和目标文件
                source_file = os.path.join(source_dir, old_filename)
                output_file = os.path.join(output_dir, new_filename)
                
                # 复制文件
                if os.path.exists(source_file):
                    import shutil
                    shutil.copy2(source_file, output_file)
                    graph_files_created.append(new_filename)
                    self.paint_log(f"已创建: {new_filename}")
                else:
                    self.paint_log(f"警告: 源文件不存在 - {source_file}")
                    continue
                
                # 生成新的模型动作编号 - 与原车逻辑一致
                if len(model_rows) == 0:
                    # 第一个：基于max_model_id递增1
                    last_unit = max_model_id % 10
                    prefix = max_model_id // 100
                    
                    if last_unit == 9:
                        new_unit = 0
                        new_prefix = prefix + 1
                    else:
                        new_unit = last_unit + 1
                        new_prefix = prefix
                    
                    new_model_id = new_prefix * 100 + new_unit
                    
                    if new_model_id < 10000:
                        new_model_id = 26010
                else:
                    # 后续：在前一行基础上+10
                    new_model_id = max_model_id + 10
                
                # 检查冲突
                counter = 0
                while new_model_id in df_model['编号'].values and counter < 1000:
                    new_model_id += 10
                    counter += 1
                
                max_model_id = new_model_id
                
                # 生成新备注：使用用户输入的模型备注，在第一个下划线后插入涂装名称
                # 例如: template-base-show -> template_涂装名-base-show
                # 如果用户没有输入模型备注，则使用原车的备注
                base_remark = model_remark if model_remark else (str(old_remark) if old_remark else "")
                remark_parts = base_remark.split('_', 1)
                if len(remark_parts) >= 2:
                    new_remark = f"{remark_parts[0]}_{paint_name}_{remark_parts[1]}"
                else:
                    new_remark = f"{base_remark}_{paint_name}" if base_remark else paint_name
                
                # 根据新文件名判断车身/底盘动作
                # graph名字中带base的填在车身动作列，带chassis的填在底盘动作列
                filename_lower = new_filename.lower()
                if 'base' in filename_lower:
                    body_action = new_model_id  # 车身动作填入模型动作编号
                    chassis_action = ''  # 底盘动作留空
                elif 'chassis' in filename_lower:
                    body_action = ''  # 车身动作留空
                    chassis_action = new_model_id  # 底盘动作填入模型动作编号
                else:
                    # 如果都没有，默认填入车身动作
                    body_action = new_model_id
                    chassis_action = ''
                
                # 获取原记录中的事件信息
                start_event = record.get('开始事件', '')
                end_event = record.get('结束事件', '')
                reenter_event = record.get('重入事件', '')
                force_end_event = record.get('强制结束事件', '')
                
                # 新graph路径
                new_graph_path = f"Char\\\\graph\\\\car_parts\\\\{new_filename}"
                
                # 添加模型动作记录
                model_rows.append({
                    'skip': '要',
                    '编号': new_model_id,
                    '备注': new_remark,
                    'GRAPH路径': new_graph_path,
                    '精简GRAPH路径': '',
                    '开始事件': start_event,
                    '重入事件': reenter_event,
                    '结束事件': end_event,
                    '强制结束事件': force_end_event,
                    '状态变量名1': '',
                    '状态变量值1': '',
                    '状态变量名2': '',
                    '状态变量值2': '',
                    '状态变量名3': '',
                    '状态变量值3': '',
                    '状态变量名4': '',
                    '状态变量值4': '',
                    '状态变量名5': '',
                    '状态变量值5': '',
                    '标志变量名': '',
                    '标志变量值': '',
                    '阶段1事件': '',
                    '阶段2事件': '',
                    '阶段结束事件': '',
                    '叠加次数变量名': ''
                })
                
                # 查找对应的车辆动作记录
                old_model_id = record['编号']
                old_vehicle_records = df_vehicle[df_vehicle['动作编号'] == old_model_id]
                
                for v_idx, v_record in old_vehicle_records.iterrows():
                    # 新车辆动作编号
                    vehicle_id += 1
                    
                    # 保留原老车的车辆动作设置
                    old_is_hit = v_record.get('是否受击', '')
                    old_vehicle_type = v_record.get('车辆类型', '')
                    
                    # 添加车辆动作记录
                    vehicle_rows.append({
                        'skip': '要',
                        '编号': vehicle_id,
                        '动作编号': new_model_id,
                        '模型': str(model_number),
                        '是否受击': old_is_hit,
                        '车辆类型': old_vehicle_type,
                        '备注': new_remark,
                        '车身动作': body_action,
                        '底盘动作': chassis_action,
                        '附加模型动作': v_record.get('附加模型动作', ''),
                        '音效资源': v_record.get('音效资源', ''),
                        '特殊效果': v_record.get('特殊效果', ''),
                        '挂接编号': v_record.get('挂接编号', ''),
                        '排气管特效编号': v_record.get('排气管特效编号', ''),
                        '排气管挂接方案': v_record.get('排气管挂接方案', ''),
                        '放置友方模型': v_record.get('放置友方模型', ''),
                        '放置自身模型': v_record.get('放置自身模型', ''),
                        '放置敌方模型': v_record.get('放置敌方模型', ''),
                        '挂接模型动作1': v_record.get('挂接模型动作1', ''),
                        '挂接模型动作2': v_record.get('挂接模型动作2', ''),
                        '挂接模型动作3': v_record.get('挂接模型动作3', ''),
                        '车身脚本挂接特效': v_record.get('车身脚本挂接特效', ''),
                        '底盘脚本挂接特效': v_record.get('底盘脚本挂接特效', ''),
                        '角色动作': v_record.get('角色动作', '')
                    })
            
            # 保存到Excel
            output_excel = r"E:\H65\Doc\X_导入数据\01_美术资源\模型动作.xlsx"
            wb = load_workbook(output_excel)
            
            # 模型动作
            if model_rows:
                ws_model = wb['模型动作']
                for row_data in model_rows:
                    row = []
                    for col in ['skip', '编号', '备注', 'GRAPH路径', '精简GRAPH路径', '开始事件', 
                               '重入事件', '结束事件', '强制结束事件', '状态变量名1', '状态变量值1', 
                               '状态变量名2', '状态变量值2', '状态变量名3', '状态变量值3', 
                               '状态变量名4', '状态变量值4', '状态变量名5', '状态变量值5', 
                               '标志变量名', '标志变量值', '阶段1事件', '阶段2事件', 
                               '阶段结束事件', '叠加次数变量名']:
                        val = row_data.get(col, '')
                        if val == '':
                            val = None
                        row.append(val)
                    ws_model.append(row)
                self.paint_log(f"已添加 {len(model_rows)} 条模型动作记录")
            
            # 车辆动作
            if vehicle_rows:
                ws_vehicle = wb['车辆动作']
                for row_data in vehicle_rows:
                    row = []
                    for col in ['skip', '编号', '动作编号', '模型', '是否受击', '车辆类型', '备注', 
                               '车身动作', '底盘动作', '附加模型动作', '音效资源', '特殊效果', 
                               '挂接编号', '排气管特效编号', '排气管挂接方案', '放置友方模型', 
                               '放置自身模型', '放置敌方模型', '挂接模型动作1', '挂接模型动作2', 
                               '挂接模型动作3', '车身脚本挂接特效', '底盘脚本挂接特效', '角色动作']:
                        val = row_data.get(col, '')
                        if val == '':
                            val = None
                        row.append(val)
                    ws_vehicle.append(row)
                self.paint_log(f"已添加 {len(vehicle_rows)} 条车辆动作记录")
            
            wb.save(output_excel)
            self.paint_log(f"涂装生成完成！")
            
            messagebox.showinfo("成功", f"已生成 {len(graph_files_created)} 个涂装文件并填表")
            
        except Exception as e:
            self.paint_log(f"错误: {str(e)}")
            import traceback
            traceback.print_exc()
            messagebox.showerror("错误", f"生成失败:\n{str(e)}")
    
    def create_preview_tab(self):
        """创建预览标签页"""
        # 预览文本
        preview_frame = ttk.LabelFrame(self.preview_frame, text=" XML预览 ", padding="10")
        preview_frame.pack(fill=tk.BOTH, expand=True)
        
        text_frame = ttk.Frame(preview_frame)
        text_frame.pack(fill=tk.BOTH, expand=True)
        
        self.preview_text = tk.Text(text_frame, wrap=tk.NONE, font=("Consolas", 9))
        self.preview_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        v_scroll = ttk.Scrollbar(text_frame, orient=tk.VERTICAL, command=self.preview_text.yview)
        v_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        self.preview_text.config(yscrollcommand=v_scroll.set)
        
        h_scroll = ttk.Scrollbar(preview_frame, orient=tk.HORIZONTAL, command=self.preview_text.xview)
        h_scroll.pack(fill=tk.X)
        self.preview_text.config(xscrollcommand=h_scroll.set)
        
        # 刷新按钮
        ttk.Button(preview_frame, text="刷新预览", command=self.refresh_preview).pack(pady=10)
    
    def create_status_bar(self):
        """创建状态栏"""
        self.status_bar = ttk.Label(self.root, text="就绪", relief=tk.SUNKEN, anchor=tk.W)
        self.status_bar.pack(side=tk.BOTTOM, fill=tk.X)
    
    # ==================== 事件处理 ====================
    
    def log(self, message):
        """添加日志"""
        try:
            self.log_text.config(state='normal')
            self.log_text.insert(tk.END, message + "\n")
            self.log_text.see(tk.END)
            self.log_text.config(state='disabled')
        except:
            pass
        
        # 同时更新状态栏
        self.status_bar.config(text=message)
    
    def on_graph_name_change(self, event):
        """Graph名称更改"""
        self.model.graph_name = self.entry_graph_name.get()
    
    def refresh_layer_list(self):
        """刷新技能列表"""
        self.layer_listbox.delete(0, tk.END)
        for i, layer in enumerate(self.model.layers):
            type_str = "三段式" if layer.structure_type == StructureType.THREE_STAGE else "瞬发"
            self.layer_listbox.insert(tk.END, f"{layer.name} ({type_str})")
        
        if self.model.layers:
            self.layer_listbox.selection_set(self.current_layer_index)
            self.layer_listbox.see(self.current_layer_index)
        
        # 强制刷新UI
        self.root.update_idletasks()
    
    def on_layer_select(self, event):
        """技能选择事件"""
        selection = self.layer_listbox.curselection()
        if selection:
            index = selection[0]
            self.load_layer_to_ui(index)
    
    def load_layer_to_ui(self, index):
        """加载技能到UI"""
        if not (0 <= index < len(self.model.layers)):
            return
        
        self.current_layer_index = index
        layer = self.model.layers[index]
        
        # 更新技能名称
        self.entry_skill_name.delete(0, tk.END)
        self.entry_skill_name.insert(0, layer.name)
        
        # 更新结构类型
        if layer.structure_type == StructureType.THREE_STAGE:
            self.structure_combo.current(0)
        else:
            self.structure_combo.current(1)
        
        # 更新车辆类型
        try:
            self.var_car_type.trace_vdelete('w', self.trace_car_type_id)
        except:
            pass
        # 获取car_type的枚举名称（NEW_CAR或OLD_CAR）
        car_type_val = layer.car_type.name if hasattr(layer.car_type, 'name') else str(layer.car_type)
        self.var_car_type.set(car_type_val)
        self.trace_car_type_id = self.var_car_type.trace('w', self.on_place_change)
        
        # 更新场所设置 - 先移除trace避免触发on_place_change
        try:
            self.var_showroom_body.trace_vdelete('w', self.trace_showroom_id)
            self.var_showroom_chassis.trace_vdelete('w', self.trace_showroom_chassis_id)
            self.var_race_body.trace_vdelete('w', self.trace_race_id)
            self.var_race_chassis.trace_vdelete('w', self.trace_race_chassis_id)
        except:
            pass
        
        # 根据layer的设置确定展厅和赛道的复选框状态
        self.var_showroom_body.set(layer.enable_showroom_body)
        self.var_showroom_chassis.set(layer.enable_showroom_chassis)
        self.var_race_body.set(layer.enable_race_body)
        self.var_race_chassis.set(layer.enable_race_chassis)
        
        # 恢复trace
        self.trace_showroom_id = self.var_showroom_body.trace('w', self.on_place_change)
        self.trace_showroom_chassis_id = self.var_showroom_chassis.trace('w', self.on_place_change)
        self.trace_race_id = self.var_race_body.trace('w', self.on_place_change)
        self.trace_race_chassis_id = self.var_race_chassis.trace('w', self.on_place_change)
        
        # 更新节点下拉框
        self.update_node_combo()
        
        self.log(f"已加载技能: {layer.name}")
    
    def update_node_combo(self):
        """更新节点下拉框"""
        if not (0 <= self.current_layer_index < len(self.model.layers)):
            return
        
        layer = self.model.layers[self.current_layer_index]
        node_names = []
        
        if layer.structure_type == StructureType.THREE_STAGE:
            # 三段式节点
            nodes_map = {
                "start (开始)": layer.start_node,
                "loop01 (循环)": layer.loop_node,
                "end (结束)": layer.end_node,
                "BlendTree_#1": layer.blend_tree
            }
            for name, node in nodes_map.items():
                if node:
                    node_names.append(name)
        else:
            # 瞬发式节点
            nodes_map = {
                "BlendTree": layer.instant_blend_tree,
                "no_anim": layer.instant_action_nodes[0] if len(layer.instant_action_nodes) > 0 else None,
                "no_anim_#1": layer.instant_action_nodes[1] if len(layer.instant_action_nodes) > 1 else None
            }
            for name, node in nodes_map.items():
                if node:
                    node_names.append(name)
        
        self.node_combo['values'] = node_names
        if node_names:
            self.node_combo.current(0)
            self.on_node_select(None)
    
    def on_node_select(self, event):
        """节点选择事件"""
        selection = self.node_combo.current()
        if selection < 0:
            return
        
        if not (0 <= self.current_layer_index < len(self.model.layers)):
            return
        
        layer = self.model.layers[self.current_layer_index]
        
        # 获取选中的节点
        if layer.structure_type == StructureType.THREE_STAGE:
            nodes = [layer.start_node, layer.loop_node, layer.end_node, layer.blend_tree]
        else:
            nodes = [layer.instant_blend_tree] + layer.instant_action_nodes
        
        if selection < len(nodes):
            node = nodes[selection]
            if node:
                self.load_node_to_form(node)
                self.refresh_event_list()
    
    def load_node_to_form(self, node: GraphNode):
        """加载节点到表单"""
        self.current_node = node
        
        # 更新表单值
        self.entry_anim_name.delete(0, tk.END)
        self.entry_anim_name.insert(0, node.anim_name)
        
        self.spin_add_ref.set(node.add_ref_time)
        
        # 从节点读取single_play值
        self.var_single_play.set(node.single_play)
        
        # 强制刷新UI显示
        self.node_form_frame.update_idletasks()
    
    def apply_node_changes(self):
        """应用节点更改"""
        if not hasattr(self, 'current_node') or not self.current_node:
            messagebox.showwarning("警告", "请先选择一个节点")
            return
        
        try:
            self.current_node.anim_name = self.entry_anim_name.get()
            self.current_node.add_ref_time = float(self.spin_add_ref.get())
            self.current_node.single_play = self.var_single_play.get()
            
            self.log(f"节点 {self.current_node.name} 已更新")
            messagebox.showinfo("成功", "节点属性已更新")
        except Exception as e:
            messagebox.showerror("错误", f"更新失败: {str(e)}")
    
    def on_skill_name_change(self, event):
        """技能名称更改 - 实时同步"""
        if not (0 <= self.current_layer_index < len(self.model.layers)):
            return
        
        new_name = self.entry_skill_name.get()
        self.model.layers[self.current_layer_index].name = new_name
        # 刷新技能列表显示
        self.refresh_layer_list()
        # 刷新预览
        try:
            self.test_export()
        except:
            pass
    
    def on_structure_change(self, event):
        """结构类型更改 - 直接切换无需确认"""
        if not (0 <= self.current_layer_index < len(self.model.layers)):
            return
        
        current_type = self.model.layers[self.current_layer_index].structure_type
        new_type = StructureType.THREE_STAGE if self.structure_combo.current() == 0 else StructureType.INSTANT
        
        if current_type != new_type:
            # 保存当前场所和车辆类型设置
            old_car_type = self.model.layers[self.current_layer_index].car_type
            old_showroom_body = self.model.layers[self.current_layer_index].enable_showroom_body
            old_showroom_chassis = self.model.layers[self.current_layer_index].enable_showroom_chassis
            old_race_body = self.model.layers[self.current_layer_index].enable_race_body
            old_race_chassis = self.model.layers[self.current_layer_index].enable_race_chassis
            
            # 直接重建技能，无需确认
            # 重建技能
            self.model.layers[self.current_layer_index] = GraphLayer()
            self.model.layers[self.current_layer_index].name = self.entry_skill_name.get()
            self.model.layers[self.current_layer_index].scene_pos_y = 32 + self.current_layer_index * 200
            
            # 恢复场所和车辆类型设置
            self.model.layers[self.current_layer_index].car_type = old_car_type
            self.model.layers[self.current_layer_index].enable_showroom_body = old_showroom_body
            self.model.layers[self.current_layer_index].enable_showroom_chassis = old_showroom_chassis
            self.model.layers[self.current_layer_index].enable_race_body = old_race_body
            self.model.layers[self.current_layer_index].enable_race_chassis = old_race_chassis
            
            if new_type == StructureType.THREE_STAGE:
                self.model.layers[self.current_layer_index].create_three_stage_structure()
            else:
                self.model.layers[self.current_layer_index].create_instant_structure()
            
            self.refresh_layer_list()
            self.load_layer_to_ui(self.current_layer_index)
            self.log(f"技能已转换为: {'三段式' if new_type == StructureType.THREE_STAGE else '瞬发式'}")
    
    def on_place_change(self, *args):
        """场所设置更改"""
        if not (0 <= self.current_layer_index < len(self.model.layers)):
            return
        
        layer = self.model.layers[self.current_layer_index]
        
        # 更新车辆类型
        car_type_val = self.var_car_type.get()
        layer.car_type = CarType.NEW_CAR if car_type_val == "NEW_CAR" else CarType.OLD_CAR
        
        # 更新场所类型（展厅车身/底盘、赛道车身/底盘）
        layer.enable_showroom_body = self.var_showroom_body.get()
        layer.enable_showroom_chassis = self.var_showroom_chassis.get()
        layer.enable_race_body = self.var_race_body.get()
        layer.enable_race_chassis = self.var_race_chassis.get()
    
    # ==================== 技能管理 ====================
    
    def add_layer(self):
        """添加技能"""
        # 弹出对话框让用户输入名称
        name = tk.simpledialog.askstring("添加技能", "请输入技能名称:", parent=self.root)
        if not name:
            return
        
        # 弹出对话框选择结构类型
        dialog = tk.Toplevel(self.root)
        dialog.title("选择结构类型")
        dialog.geometry("300x150")
        dialog.transient(self.root)
        dialog.grab_set()
        
        ttk.Label(dialog, text="请选择技能结构类型:").pack(pady=20)
        
        result = {"type": None}
        
        def select_type(struct_type):
            result["type"] = struct_type
            dialog.destroy()
        
        btn_frame = ttk.Frame(dialog)
        btn_frame.pack(pady=10)
        
        ttk.Button(btn_frame, text="三段式 (Start-Loop-End)", 
                   command=lambda: select_type(StructureType.THREE_STAGE)).pack(pady=5, fill=tk.X, padx=20)
        ttk.Button(btn_frame, text="瞬发式", 
                   command=lambda: select_type(StructureType.INSTANT)).pack(pady=5, fill=tk.X, padx=20)
        
        self.root.wait_window(dialog)
        
        if not result["type"]:
            return
        
        # 创建技能
        layer = self.model.add_layer(result["type"])
        layer.name = name
        layer.scene_pos_y = 32 + (len(self.model.layers) - 1) * 200
        
        self.refresh_layer_list()
        
        index = len(self.model.layers) - 1
        self.layer_listbox.selection_clear(0, tk.END)
        self.layer_listbox.selection_set(index)
        self.load_layer_to_ui(index)
        
        self.log(f"添加了技能: {name}")
    
    def delete_layer(self):
        """删除技能"""
        if len(self.model.layers) <= 1:
            messagebox.showwarning("警告", "至少需要保留一个技能")
            return
        
        reply = messagebox.askyesno("确认", "确定要删除当前技能吗?")
        if reply:
            self.model.remove_layer(self.current_layer_index)
            self.refresh_layer_list()
            
            if self.current_layer_index >= len(self.model.layers):
                self.current_layer_index = len(self.model.layers) - 1
            
            # 重新计算位置
            for i, layer in enumerate(self.model.layers):
                layer.scene_pos_y = 32 + i * 200
            
            self.load_layer_to_ui(self.current_layer_index)
            self.log("已删除技能")
    
    # ==================== 事件管理 ====================
    
    def refresh_event_list(self):
        """刷新事件列表"""
        self.event_listbox.delete(0, tk.END)
        
        if not hasattr(self, 'current_node') or not self.current_node:
            return
        
        for event in self.current_node.events:
            self.event_listbox.insert(tk.END, f"轨道:{event._TrackName} 时间:{event.TimePer} 事件:{event.Name}")
    
    def add_event(self):
        """添加事件"""
        if not hasattr(self, 'current_node') or not self.current_node:
            messagebox.showwarning("警告", "请先选择一个节点")
            return
        
        # 简单对话框获取事件名
        event_name = tk.simpledialog.askstring("添加事件", "请输入事件名称:", parent=self.root)
        if event_name:
            event = Event(_TrackName="Event01", TimePer=0, Name=event_name)
            self.current_node.events.append(event)
            self.refresh_event_list()
            self.log(f"添加了事件: {event_name}")
    
    def delete_event(self):
        """删除事件"""
        selection = self.event_listbox.curselection()
        if not selection:
            messagebox.showwarning("警告", "请先选择一个事件")
            return
        
        if not hasattr(self, 'current_node') or not self.current_node:
            return
        
        index = selection[0]
        if index < len(self.current_node.events):
            event_name = self.current_node.events[index].Name
            del self.current_node.events[index]
            self.refresh_event_list()
            self.log(f"删除了事件: {event_name}")
    
    # ==================== 导出 ====================
    
    def update_filename_preview(self, *args):
        """更新文件名预览 - 根据每个技能的场所设置"""
        template_name = self.entry_graph_name.get().strip() or "模版名称"
        
        # 收集所有技能的场所+类型设置
        export_keys = set()
        for layer in self.model.layers:
            if layer.enable_showroom_body:
                export_keys.add(('show', 'base'))
            if layer.enable_showroom_chassis:
                export_keys.add(('show', 'chassis'))
            if layer.enable_race_body:
                export_keys.add(('race', 'base'))
            if layer.enable_race_chassis:
                export_keys.add(('race', 'chassis'))
        
        if not export_keys:
            self.label_filename_preview.config(text="(请至少配置一个技能的场所类型)", foreground="#e74c3c")
            return
        
        # 生成预览文本
        key_names = []
        for (place, part) in sorted(export_keys):
            place_name = "展厅" if place == "show" else "赛道"
            part_name = "车身" if part == "base" else "底盘"
            key_names.append(f"{place_name}{part_name}")
        
        self.label_filename_preview.config(text=f"将导出: {', '.join(key_names)}.graph", foreground="#27ae60")
    
    def browse_output(self):
        """浏览输出路径"""
        filename = filedialog.askdirectory(title="选择输出目录")
        if filename:
            self.entry_output_path.delete(0, tk.END)
            self.entry_output_path.insert(0, filename)
    
    def export_graph(self):
        """导出Graph文件 - 每个技能可单独设置场所和类型"""
        # 更新Graph名称
        self.model.graph_name = self.entry_graph_name.get()
        
        # 获取输出目录
        output_dir = self.entry_output_path.get().strip()
        if not output_dir:
            output_dir = "."
        
        template_name = self.entry_graph_name.get().strip() or "output"
        
        # 获取模型信息
        model_number = self.entry_model_number.get().strip()
        model_remark = self.entry_model_remark.get().strip()
        
        try:
            # 收集需要导出的文件列表
            # 结构: {(place, part): [(layer_index, layer_name, car_type), ...]}
            # place: show/race, part: base/chassis
            export_tasks = {}
            
            for idx, layer in enumerate(self.model.layers):
                # 检查展厅设置
                if layer.enable_showroom_body:
                    key = ('show', 'base')
                    if key not in export_tasks:
                        export_tasks[key] = []
                    export_tasks[key].append((idx, layer.name, layer.car_type))
                
                if layer.enable_showroom_chassis:
                    key = ('show', 'chassis')
                    if key not in export_tasks:
                        export_tasks[key] = []
                    export_tasks[key].append((idx, layer.name, layer.car_type))
                
                # 检查赛道设置
                if layer.enable_race_body:
                    key = ('race', 'base')
                    if key not in export_tasks:
                        export_tasks[key] = []
                    export_tasks[key].append((idx, layer.name, layer.car_type))
                
                if layer.enable_race_chassis:
                    key = ('race', 'chassis')
                    if key not in export_tasks:
                        export_tasks[key] = []
                    export_tasks[key].append((idx, layer.name, layer.car_type))
            
            if not export_tasks:
                messagebox.showwarning("警告", "请至少在一个技能中配置展厅或赛道")
                return
            
            # 批量导出graph文件
            exported_files = []
            graph_data_list = []  # 存储每个graph的数据用于填表
            
            # 为每个场所+类型组合导出文件
            for (place, part), layer_info_list in sorted(export_tasks.items()):
                place_name = "展厅" if place == "show" else "赛道"
                part_name = "车身" if part == "base" else "底盘"
                
                # 创建只包含该场所技能的临时model
                temp_model = GraphModel()
                temp_model.graph_name = self.model.graph_name
                
                for idx, layer_name, car_type in layer_info_list:
                    # 复制原始layer到临时model
                    original_layer = self.model.layers[idx]
                    # 使用深拷贝
                    import copy
                    temp_layer = copy.deepcopy(original_layer)
                    temp_model.layers.append(temp_layer)
                
                # 序列化并保存
                serializer = GraphSerializer(temp_model)
                content = serializer.serialize()
                
                filename = f"{template_name}_{part}_{place}.graph"
                filepath = os.path.join(output_dir, filename)
                
                # 保存文件
                with open(filepath, 'w', encoding='utf-8') as f:
                    f.write(content)
                
                # 解析graph获取技能信息
                skill_info = self.parse_graph_events(content, template_name, part, place)
                skill_info['car_type'] = layer_info_list[0][2] if layer_info_list else CarType.NEW_CAR
                graph_data_list.append(skill_info)
                
                exported_files.append(filename)
                self.log(f"已导出: {filename} (包含 {len(layer_info_list)} 个技能)")
            
            self.log(f"批量导出完成！共导出 {len(exported_files)} 个文件")
            
            # 如果有模型编号，自动填表
            if model_number:
                try:
                    model_number = int(model_number)
                    # 传递所有场所+类型组合列表
                    all_places = list(export_tasks.keys())
                    # 传递原始layers信息用于填表
                    self.auto_fill_excel(model_number, model_remark, all_places, graph_data_list, output_dir, template_name, self.model.layers)
                except ValueError:
                    messagebox.showwarning("警告", "模型编号必须是数字，已跳过自动填表")
            
            messagebox.showinfo("成功", f"已成功导出 {len(exported_files)} 个文件:\n" + "\n".join(exported_files))
        except Exception as e:
            self.log(f"导出失败: {str(e)}")
            messagebox.showerror("错误", f"导出失败:\n{str(e)}")
    
    def parse_graph_events(self, content, template_name, part, place):
        """解析graph内容，提取技能事件信息"""
        try:
            root = ET.fromstring(content)
            
            skills = []
            
            # 查找Layer节点
            for blend_tree in root.iter('Input'):
                if blend_tree.find('Name') is not None and blend_tree.find('Type') is not None:
                    if blend_tree.find('Type').text == 'BlendTree':
                        # 查找Layer
                        for layer_input in blend_tree.findall('Input'):
                            if layer_input.find('Type') is not None and layer_input.find('Type').text == 'Layer':
                                # 遍历Layer下的每个技能StateMachine（直接子节点）
                                for skill_sm in layer_input.findall('Input'):
                                    if skill_sm.find('Type') is not None and skill_sm.find('Type').text == 'StateMachine':
                                        skill_name = skill_sm.find('Name').text if skill_sm.find('Name') is not None else ""
                                        
                                        # 判断结构类型 - 查找该StateMachine下的ActionNode
                                        has_no_anim = False
                                        has_end = False
                                        
                                        for action_node in skill_sm.findall('Input'):
                                            if action_node.find('Type') is not None and action_node.find('Type').text == 'ActionNode':
                                                node_name = action_node.find('Name').text if action_node.find('Name') is not None else ""
                                                if node_name in ['no_anim', 'no_anim_#1']:
                                                    has_no_anim = True
                                                if node_name == 'end':
                                                    has_end = True
                                        
                                        if has_no_anim:
                                            current_skill_type = "瞬发式"
                                            target_node = 'no_anim'
                                        elif has_end:
                                            current_skill_type = "三段式"
                                            target_node = 'start'
                                        else:
                                            current_skill_type = "三段式"
                                            target_node = 'start'
                                        
                                        # 获取代表节点的事件
                                        for action_node in skill_sm.findall('Input'):
                                            if action_node.find('Type') is not None and action_node.find('Type').text == 'ActionNode':
                                                node_name = action_node.find('Name').text if action_node.find('Name') is not None else ""
                                                
                                                if node_name == target_node:
                                                    events = []
                                                    for event in action_node.findall('Event'):
                                                        if event.find('Name') is not None:
                                                            events.append(event.find('Name').text)
                                                    
                                                    single_play_elem = action_node.find('SinglePlay')
                                                    single_play = single_play_elem.text if single_play_elem is not None else "true"
                                                    
                                                    skills.append({
                                                        'name': skill_name,
                                                        'events': events,
                                                        'single_play': single_play,
                                                        'skill_type': current_skill_type
                                                    })
                                                    break
            
            return {
                'template_name': template_name,
                'part': part,
                'place': place,
                'skills': skills,
                'structure_type': current_skill_type if skills else '三段式'
            }
        except Exception as e:
            self.log(f"解析graph失败: {str(e)}")
            return {'template_name': template_name, 'part': part, 'place': place, 'skills': [], 'structure_type': '三段式'}
    
    def auto_fill_excel(self, model_number, model_remark, places, graph_data_list, output_dir, template_name, layers=None):
        """自动填表到Excel - 填写三个页签
        places: 场所+类型组合列表，如 [('show', 'base'), ('race', 'chassis')]
        layers: 包含所有技能层的列表，每个layer有name和places属性
        """
        try:
            # 使用layers直接构建skills_by_place（更可靠）
            # 结构: {(place, part): [{'name': xxx, 'skill_type': xxx, 'car_type': xxx}, ...]}
            skills_by_key = {}
            
            if layers:
                # 从layers直接获取技能信息和场所选择
                for layer in layers:
                    skill_name = layer.name
                    skill_type_str = "三段式" if layer.structure_type == StructureType.THREE_STAGE else "瞬发式"
                    car_type_str = "新车" if layer.car_type == CarType.NEW_CAR else "老车"
                    
                    # 根据设置确定场所+类型
                    if layer.enable_showroom_body:
                        key = ('show', 'base')
                        if key not in skills_by_key:
                            skills_by_key[key] = []
                        skills_by_key[key].append({
                            'name': skill_name,
                            'skill_type': skill_type_str,
                            'car_type': layer.car_type
                        })
                    
                    if layer.enable_showroom_chassis:
                        key = ('show', 'chassis')
                        if key not in skills_by_key:
                            skills_by_key[key] = []
                        skills_by_key[key].append({
                            'name': skill_name,
                            'skill_type': skill_type_str,
                            'car_type': layer.car_type
                        })
                    
                    if layer.enable_race_body:
                        key = ('race', 'base')
                        if key not in skills_by_key:
                            skills_by_key[key] = []
                        skills_by_key[key].append({
                            'name': skill_name,
                            'skill_type': skill_type_str,
                            'car_type': layer.car_type
                        })
                    
                    if layer.enable_race_chassis:
                        key = ('race', 'chassis')
                        if key not in skills_by_key:
                            skills_by_key[key] = []
                        skills_by_key[key].append({
                            'name': skill_name,
                            'skill_type': skill_type_str,
                            'car_type': layer.car_type
                        })
                
                self.log(f"直接从layers构建skills_by_key: {skills_by_key}")
            else:
                # 旧逻辑：从graph_data_list构建（备用）
                for graph_data in graph_data_list:
                    place = graph_data['place']
                    skills_list = graph_data.get('skills', [])
                    self.log(f"场所: {place}, 技能数: {len(skills_list)}")
                    
                    if place not in skills_by_place:
                        skills_by_key[place] = []
                    
                    skills_by_key[place].extend(skills_list)
            
            # 读取Excel
            excel_path = r"E:\H65\Doc\X_导入数据\01_美术资源\模型动作.xlsx"
            xl = pd.ExcelFile(excel_path)
            df_action = pd.read_excel(xl, '动作组合')  # 动作组合
            df_vehicle = pd.read_excel(xl, '车辆动作')  # 车辆动作
            df_model = pd.read_excel(xl, '模型动作')  # 模型动作
            
            # 查找现有最大编号
            max_action_id = df_action['编号'].max() if len(df_action) > 0 else 8000
            max_vehicle_id = df_vehicle['编号'].max() if len(df_vehicle) > 0 else 0
            
            # 查找模型动作中最后一行的编号（只找2开头的5位数）
            df_model['编号_str'] = df_model['编号'].astype(str)
            df_model_5digit = df_model[df_model['编号_str'].str.len() == 5]
            # 只取2开头的编号（26000-29999）
            df_model_2xxx = df_model_5digit[df_model_5digit['编号'] >= 26000]
            # 取最后一行的编号
            max_model_id = int(df_model_2xxx['编号'].iloc[-1]) if len(df_model_2xxx) > 0 else 26000
            
            if pd.isna(max_action_id):
                max_action_id = 8000
            if pd.isna(max_vehicle_id):
                max_vehicle_id = 0
            if pd.isna(max_model_id) or max_model_id == 0:
                max_model_id = 26000
            
            # 准备数据
            action_rows = []  # 动作组合
            vehicle_rows = []  # 车辆动作
            model_rows = []  # 模型动作
            
            # 车辆动作编号（顺序递增）
            vehicle_id = int(max_vehicle_id) + 1 if max_vehicle_id and max_vehicle_id > 0 else 1
            
            # 收集所有模型动作编号（用于车辆动作引用）
            model_ids_by_key = {}
            
            self.log(f"skills_by_key: {skills_by_key}")
            
            # 定义排序顺序：展厅车身→展厅底盘→赛道车身→赛道底盘
            def sort_key(item):
                place, part = item
                place_order = {'show': 0, 'race': 1}
                part_order = {'base': 0, 'chassis': 1}
                return (place_order.get(place, 0), part_order.get(part, 0))
            
            # 按场所+类型顺序处理
            for (place, part) in sorted(skills_by_key.keys(), key=sort_key):
                place_name = "展厅" if place == "show" else "赛道"
                part_name = "车身" if part == "base" else "底盘"
                vehicle_type_text = part_name
                
                all_skills = skills_by_key[(place, part)]
                
                if not all_skills:
                    continue
                
                # 按结构类型排序：三段式在前，瞬发在后
                sorted_skills = sorted(all_skills, key=lambda s: 0 if s.get('skill_type') == '三段式' else 1)
                
                # 该place内的序号（从1开始）
                three_stage_seq = 1
                instant_seq = 1
                
                # 处理该place的每个技能
                for skill in sorted_skills:
                    skill_type = skill.get('skill_type', '三段式')
                    
                    # 构建备注：模型备注-技能名称-大技能xxx-展厅/赛道 类型
                    skill_name = skill.get('name', '')
                    remark = f"{model_remark or template_name}-{skill_name}-大技能{vehicle_type_text}({part})-{place_name} {skill_type}"
                    
                    # 计算模型动作编号 - 递增逻辑
                    if len(model_rows) == 0:
                        # 第一个skill
                        last_unit = max_model_id % 10
                        prefix = max_model_id // 100
                        
                        if last_unit == 9:
                            new_unit = 0
                            new_prefix = prefix + 1
                        else:
                            new_unit = last_unit + 1
                            new_prefix = prefix
                        
                        model_id = new_prefix * 100 + new_unit
                        
                        if model_id < 10000:
                            model_id = 26010
                    else:
                        # 后续skill：在前一行基础上+10
                        model_id = max_model_id + 10
                    
                    # 检查冲突
                    counter = 0
                    while model_id in df_model['编号'].values and counter < 1000:
                        model_id += 10
                        counter += 1
                    
                    max_model_id = model_id
                    
                    # 生成事件名称 - 每个graph独立编号
                    if skill_type == "瞬发式":
                        seq_str = f"{instant_seq:02d}"
                        start_event = f"add_power_start_{seq_str}"
                        reenter_event = f"add_power_start_{seq_str}"
                        end_event = f"add_power_stop_{seq_str}"
                        force_end_event = f"add_power_stop_{seq_str}"
                        instant_seq += 1
                    else:
                        seq_str = f"{three_stage_seq:02d}"
                        start_event = f"BigSkill_Start_{seq_str}"
                        reenter_event = f"BigSkill_Start_{seq_str}"
                        end_event = f"BigSkill_Stop_{seq_str}"
                        force_end_event = f"BigSkill_Stop_{seq_str}"
                        three_stage_seq += 1
                    
                    # 保存每个场所+类型对应的模型动作编号
                    key = (place, part)
                    if key not in model_ids_by_key:
                        model_ids_by_key[key] = []
                    model_ids_by_key[key].append(model_id)
                    
                    # 构建GRAPH路径
                    graph_path = f"Char\\graph\\car_parts\\{template_name}_{part}_{place}.graph"
                    
                    # 添加模型动作记录
                    model_rows.append({
                        'skip': '要',
                        '编号': model_id,
                        '备注': remark,
                        'GRAPH路径': graph_path,
                        '精简GRAPH路径': '',
                        '开始事件': start_event,
                        '重入事件': reenter_event,
                        '结束事件': end_event,
                        '强制结束事件': force_end_event,
                        '状态变量名1': '',
                        '状态变量值1': '',
                        '状态变量名2': '',
                        '状态变量值2': '',
                        '状态变量名3': '',
                        '状态变量值3': '',
                        '状态变量名4': '',
                        '状态变量值4': '',
                        '状态变量名5': '',
                        '状态变量值5': '',
                        '标志变量名': '',
                        '标志变量值': '',
                        '阶段1事件': '',
                        '阶段2事件': '',
                        '阶段结束事件': '',
                        '叠加次数变量名': ''
                    })
            
            # 添加动作组合记录
            action_rows.append({
                '编号': model_number,
                '备注': model_remark or template_name
            })
            
            # 为每个场所+类型添加车辆动作记录（顺序与模型动作一致）
            for (place, part) in sorted(model_ids_by_key.keys(), key=sort_key):
                place_name = "展厅" if place == "show" else "赛道"
                part_name = "车身" if part == "base" else "底盘"
                vehicle_type_text = part_name
                
                ids = model_ids_by_key[(place, part)]
                
                # 为该场所的每个技能(模型动作)创建一条车辆动作记录
                for idx, model_id in enumerate(ids):
                    # 获取对应的技能信息来构建备注
                    skill_info = skills_by_key[(place, part)][idx] if idx < len(skills_by_key.get((place, part), [])) else {}
                    skill_type = skill_info.get('skill_type', '三段式')
                    skill_name = skill_info.get('name', '')
                    car_type = skill_info.get('car_type', CarType.NEW_CAR)
                    
                    # 备注与模型动作页签一致：模型备注-技能名称-大技能xxx-展厅/赛道 类型
                    remark = f"{model_remark or template_name}-{skill_name}-大技能{vehicle_type_text}({part})-{place_name} {skill_type}"
                    
                    # 是否受击：根据车辆类型（新车填1，老车不填）
                    is_hit = 1 if car_type == CarType.NEW_CAR else ''
                    
                    vehicle_rows.append({
                        'skip': '要',
                        '编号': vehicle_id,  # 顺序递增
                        '动作编号': model_id,  # 与模型动作编号一致
                        '模型': str(model_number),
                        '是否受击': is_hit,  # 新车填1，老车空
                        '车辆类型': 1,  # 填1
                        '备注': remark,  # 与模型动作备注一致
                        '车身动作': model_id if part == 'base' else '',  # 车身时填
                        '底盘动作': model_id if part == 'chassis' else '',  # 底盘时填
                        '附加模型动作': '',
                        '音效资源': '',
                        '特殊效果': '',
                        '挂接编号': '',
                        '排气管特效编号': '',
                        '排气管挂接方案': '',
                        '放置友方模型': '',
                        '放置自身模型': '',
                        '放置敌方模型': '',
                        '挂接模型动作1': '',
                        '挂接模型动作2': '',
                        '挂接模型动作3': '',
                        '车身脚本挂接特效': '',
                        '底盘脚本挂接特效': '',
                        '角色动作': ''
                    })
                    vehicle_id += 1  # 递增编号
            
            # 保存到原始Excel文件（保留格式）
            output_excel = r"E:\H65\Doc\X_导入数据\01_美术资源\模型动作.xlsx"
            
            # 使用openpyxl加载并保留格式
            wb = load_workbook(output_excel)
            
            # 动作组合 - 追加数据
            if action_rows:
                ws_action = wb['动作组合']
                start_row = ws_action.max_row + 1
                for row_data in action_rows:
                    ws_action.append([row_data['编号'], row_data['备注']])
            
            # 车辆动作 - 追加数据
            if vehicle_rows:
                ws_vehicle = wb['车辆动作']
                for row_data in vehicle_rows:
                    row = []
                    for col in ['skip', '编号', '动作编号', '模型', '是否受击', '车辆类型', '备注', 
                               '车身动作', '底盘动作', '附加模型动作', '音效资源', '特殊效果', 
                               '挂接编号', '排气管特效编号', '排气管挂接方案', '放置友方模型', 
                               '放置自身模型', '放置敌方模型', '挂接模型动作1', '挂接模型动作2', 
                               '挂接模型动作3', '车身脚本挂接特效', '底盘脚本挂接特效', '角色动作']:
                        val = row_data.get(col, '')
                        # 如果值为空，设置为None以清除格式
                        if val == '':
                            val = None
                        row.append(val)
                    ws_vehicle.append(row)
            
            # 模型动作 - 追加数据
            if model_rows:
                ws_model = wb['模型动作']
                for row_data in model_rows:
                    row = []
                    for col in ['skip', '编号', '备注', 'GRAPH路径', '精简GRAPH路径', '开始事件', 
                               '重入事件', '结束事件', '强制结束事件', '状态变量名1', '状态变量值1', 
                               '状态变量名2', '状态变量值2', '状态变量名3', '状态变量值3', 
                               '状态变量名4', '状态变量值4', '状态变量名5', '状态变量值5', 
                               '标志变量名', '标志变量值', '阶段1事件', '阶段2事件', 
                               '阶段结束事件', '叠加次数变量名']:
                        row.append(row_data.get(col, ''))
                    ws_model.append(row)
            
            # 保存（保留原始格式）
            wb.save(output_excel)
            
            self.log(f"自动填表完成: {output_excel}")
            self.log(f"动作组合: {len(action_rows)}条, 车辆动作: {len(vehicle_rows)}条, 模型动作: {len(model_rows)}条")
            
        except Exception as e:
            self.log(f"自动填表失败: {str(e)}")
            messagebox.showwarning("警告", f"自动填表失败:\n{str(e)}")
    
    def test_export(self):
        """测试导出"""
        # 更新Graph名称
        self.model.graph_name = self.entry_graph_name.get()
        
        try:
            serializer = GraphSerializer(self.model)
            content = serializer.serialize()
            
            self.preview_text.config(state='normal')
            self.preview_text.delete(1.0, tk.END)
            self.preview_text.insert(1.0, content)
            self.preview_text.config(state='disabled')
            
            self.log("预览已更新")
        except Exception as e:
            self.log(f"生成预览失败: {str(e)}")
            messagebox.showerror("错误", f"生成预览失败:\n{str(e)}")
    
    def refresh_preview(self):
        """刷新预览"""
        self.test_export()
    
    # ==================== 自动填表功能 ====================
    
    def query_model_actions(self):
        """查询模型动作数据"""
        model_id = self.entry_model_id.get().strip()
        
        if not model_id:
            messagebox.showwarning("警告", "请输入模型编号")
            return
        
        try:
            model_id = int(model_id)
        except ValueError:
            messagebox.showerror("错误", "模型编号必须是数字")
            return
        
        # 读取Excel数据
        excel_path = r"E:\H65\Doc\X_导入数据\01_美术资源\模型动作.xlsx"
        
        try:
            xl = pd.ExcelFile(excel_path)
            df_action = pd.read_excel(xl, '动作组合')
            df_vehicle = pd.read_excel(xl, '车辆动作')
            df_model = pd.read_excel(xl, '模型动作')
        except Exception as e:
            messagebox.showerror("错误", f"读取Excel失败:\n{str(e)}")
            return
        
        # 查找动作组合
        action_combo = df_action[df_action['编号'] == model_id]
        if action_combo.empty:
            messagebox.showwarning("警告", f"未找到编号为 {model_id} 的动作组合")
            return
        
        model_name = action_combo.iloc[0]['备注']
        
        # 获取当前选择的车辆类型和场所
        vehicle_type = self.var_vehicle_type.get()  # base 或 chassis
        places = []
        if self.var_showroom.get():
            places.append("展厅")
        if self.var_race.get():
            places.append("赛道")
        
        # 查找对应的车辆动作
        # 车辆动作的"模型"字段包含模型编号
        matching_vehicles = df_vehicle[df_vehicle['模型'].fillna('').astype(str).str.contains(str(model_id))]
        
        if matching_vehicles.empty:
            messagebox.showwarning("警告", f"未找到包含模型 {model_id} 的车辆动作")
            return
        
        # 根据车辆类型筛选
        if vehicle_type == "base":
            action_ids = matching_vehicles['车身动作'].dropna().unique()
        else:
            action_ids = matching_vehicles['底盘动作'].dropna().unique()
        
        if len(action_ids) == 0:
            messagebox.showwarning("警告", f"未找到模型 {model_id} 对应的{('车身' if vehicle_type == 'base' else '底盘')}动作")
            return
        
        # 查找对应的模型动作
        results = []
        for action_id in action_ids[:10]:  # 限制显示数量
            action_id = int(action_id)
            model_actions = df_model[df_model['编号'] == action_id]
            
            if not model_actions.empty:
                for _, row in model_actions.iterrows():
                    remark = str(row['备注']) if pd.notna(row['备注']) else ''
                    # 根据场所筛选
                    for place in places:
                        if place in remark or (place == '展厅' and 'show' in remark.lower()) or (place == '赛道' and 'race' in remark.lower()):
                            results.append({
                                '动作编号': action_id,
                                '备注': remark,
                                'GRAPH路径': row['GRAPH路径'] if pd.notna(row['GRAPH路径']) else '',
                                '开始事件': row['开始事件'] if pd.notna(row['开始事件']) else '',
                                '结束事件': row['结束事件'] if pd.notna(row['结束事件']) else ''
                            })
        
        # 显示结果
        self.result_text.delete(1.0, tk.END)
        self.result_text.insert(tk.END, f"模型编号: {model_id}\n")
        self.result_text.insert(tk.END, f"模型名称: {model_name}\n")
        self.result_text.insert(tk.END, f"车辆类型: {'车身' if vehicle_type == 'base' else '底盘'}\n")
        self.result_text.insert(tk.END, f"场所: {', '.join(places)}\n")
        self.result_text.insert(tk.END, f"找到 {len(results)} 条动作记录:\n")
        self.result_text.insert(tk.END, "-" * 60 + "\n")
        
        for i, r in enumerate(results, 1):
            self.result_text.insert(tk.END, f"{i}. 动作编号: {r['动作编号']}\n")
            self.result_text.insert(tk.END, f"   备注: {r['备注']}\n")
            self.result_text.insert(tk.END, f"   GRAPH: {r['GRAPH路径']}\n")
            self.result_text.insert(tk.END, f"   事件: {r['开始事件']} -> {r['结束事件']}\n")
            self.result_text.insert(tk.END, "-" * 40 + "\n")
        
        self.log(f"查询完成: 模型 {model_id} ({model_name}), 找到 {len(results)} 条记录")
        
        # 保存查询结果供导出使用
        self.last_query_result = {
            'model_id': model_id,
            'model_name': model_name,
            'vehicle_type': vehicle_type,
            'places': places,
            'results': results
        }
    
    def export_to_excel(self):
        """导出到Excel"""
        if not hasattr(self, 'last_query_result') or not self.last_query_result['results']:
            messagebox.showwarning("警告", "请先查询动作数据")
            return
        
        # 选择保存路径
        save_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel文件", "*.xlsx"), ("所有文件", "*.*")],
            initialfile=f"{self.last_query_result['model_name']}_动作数据.xlsx"
        )
        
        if not save_path:
            return
        
        try:
            # 创建DataFrame
            data = []
            for r in self.last_query_result['results']:
                data.append({
                    '模型编号': self.last_query_result['model_id'],
                    '模型名称': self.last_query_result['model_name'],
                    '车辆类型': self.last_query_result['vehicle_type'],
                    '动作编号': r['动作编号'],
                    '备注': r['备注'],
                    'GRAPH路径': r['GRAPH路径'],
                    '开始事件': r['开始事件'],
                    '结束事件': r['结束事件']
                })
            
            df = pd.DataFrame(data)
            
            # 保存到Excel
            df.to_excel(save_path, index=False, engine='openpyxl')
            
            self.log(f"已导出到: {save_path}")
            messagebox.showinfo("成功", f"数据已导出到:\n{save_path}")
        except Exception as e:
            messagebox.showerror("错误", f"导出失败:\n{str(e)}")


def main():
    """主函数"""
    root = tk.Tk()
    app = GraphGeneratorApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
